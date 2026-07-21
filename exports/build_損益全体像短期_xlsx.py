import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "損益全体像（短期）"

YEARS = ["2023/5期","2024/5期","2025/5期","2026/5期","2027/5期","2028/5期","2029/5期","2030/5期","2031/5期"]
NCOL = len(YEARS)

D = {
 "売上高":        [4,7,1,0,0,0,0,1965,7470],
 "機体販売":      [0,0,0,0,0,0,0,1800,6000],
 "メンテナンス":  [0,0,0,0,0,0,0,165,1470],
 "移動サービス":  [0,0,0,0,0,0,0,0,0],
 "売上原価":      [87,175,79,0,0,0,0,720,2400],
 "人件費":        [112,121,114,258,519,649,757,1010,1270],
 "外注費":        [0,3,0,43,77,126,207,339,555],
 "地代家賃":      [26,29,17,-9,-16,-28,-39,-43,-49],
 "減価償却費":    [6,0,1,0,18,50,175,456,887],
 "研究開発費":    [0,0,150,150,200,2500,3000,3500,4000],
 "その他":        [46,59,108,40,54,76,90,102,115],
 "経常利益":      [-50,-126,-95,-139,-611,89,-4190,-4118,-1709],
 "売上高経常利益率": [-11.90,-19.14,-182.43,None,None,None,None,-2.10,-0.23],
 "営業CF":        [-39,-141,-83,-120,-563,-71,-5007,-5191,-2599],
 "投資CF":        [0,-79,1,-30,-600,-700,-4000,-7000,-10000],
 "財務CF":        [-14,184,-18,453,3822,8000,15000,50000,50000],
 "累計予約台数":  [1,2,2,2,4,8,32,80,128],
 "累計販売台数":  [0,0,0,0,0,0,0,6,26],
}

def F(**kw):
    base = {"name": "Arial", "size": 10}
    base.update(kw)
    return Font(**base)

NAVY = "1F4E79"; LIGHT = "DCE6F1"; YELLOW = "FFF2CC"
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
NUMFMT = '#,##0;[Red](#,##0);"-"'
PCTFMT = '#,##0%;[Red](#,##0%);"-"'

ws["A1"] = "損益全体像（短期）　2023/5期～2031/5期"
ws["A1"].font = F(size=14, bold=True, color=NAVY)
ws["A2"] = "単位：百万円（累計予約台数・累計販売台数は台、売上高経常利益率は％）"
ws["A2"].font = F(size=9, color="595959")

r = 4

def header_row(row):
    c = ws.cell(row=row, column=1, value="項目")
    c.font = F(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center"); c.border = border
    for j, y in enumerate(YEARS):
        c = ws.cell(row=row, column=2+j, value=y)
        c.font = F(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center"); c.border = border

def section(row, label):
    ws.cell(row=row, column=1, value=label)
    for j in range(1, NCOL+2):
        c = ws.cell(row=row, column=j)
        c.fill = PatternFill("solid", fgColor=LIGHT); c.border = border
        c.font = F(bold=True, color=NAVY)

def data_row(row, label, key, indent=False, bold=False, fill=None, fmt=NUMFMT):
    c = ws.cell(row=row, column=1, value=("　" + label) if indent else label)
    c.font = F(bold=bold); c.border = border
    if fill: c.fill = PatternFill("solid", fgColor=fill)
    for j, v in enumerate(D[key]):
        c = ws.cell(row=row, column=2+j, value=v if v is not None else "-")
        c.font = F(bold=bold); c.border = border; c.number_format = fmt
        if v is None: c.alignment = Alignment(horizontal="right")
        if fill: c.fill = PatternFill("solid", fgColor=fill)

header_row(r)
ROWMAP = {}

def put(label, key=None, **kw):
    global r
    r += 1
    if key is None:
        section(r, label)
    else:
        data_row(r, label, key, **kw)
        ROWMAP[key] = r

put("損益計算書（P/L）")
put("売上高", "売上高", bold=True)
put("うち機体販売", "機体販売", indent=True)
put("うちメンテナンス", "メンテナンス", indent=True)
put("うち移動サービス", "移動サービス", indent=True)
put("売上原価", "売上原価")
put("人件費", "人件費")
put("外注費", "外注費")
put("地代家賃", "地代家賃")
put("減価償却費", "減価償却費")
put("研究開発費", "研究開発費")
put("その他", "その他")
put("経常利益", "経常利益", bold=True, fill=YELLOW)
put("売上高経常利益率", "売上高経常利益率", fmt=PCTFMT)
put("キャッシュフロー（C/F）")
put("営業CF", "営業CF")
put("投資CF", "投資CF")
put("財務CF", "財務CF")
put("KPI（台数）")
put("累計予約台数", "累計予約台数", fmt='#,##0')
put("累計販売台数", "累計販売台数", fmt='#,##0')

r += 2
notes = [
 "※ 2023/5期～2025/5期の売上高には機体販売・メンテナンス・移動サービス以外のその他売上を含むため、内訳の合計と一致しません。",
 "※ 売上高経常利益率は出典シートの記載値です（売上高が0の期は「-」表示）。",
 "※ 各数値は出典シートの丸め値のため、表内の値どうしの加減算とは一致しない場合があります（助成金等の営業外収益は費用行に含まれません）。",
 "出典：Googleスプレッドシート「数値計画_2026-05-15_NoguchiKumazawa編集用」シート「損益全体像グラフ（可視化用-短期）」（2026-07-21時点）",
]
for n in notes:
    ws.cell(row=r, column=1, value=n).font = F(size=9, color="595959")
    r += 1

ws.column_dimensions["A"].width = 24
for j in range(2, NCOL+2):
    ws.column_dimensions[get_column_letter(j)].width = 11
ws.freeze_panes = "B5"
ws.sheet_view.showGridLines = False
ws.page_setup.orientation = "landscape"
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.sheet_properties.pageSetUpPr.fitToPage = True

cats = Reference(ws, min_col=2, max_col=1+NCOL, min_row=4)

bar = BarChart(); bar.type = "col"; bar.grouping = "stacked"; bar.overlap = 100
for key in ["機体販売", "メンテナンス", "移動サービス"]:
    bar.add_data(Reference(ws, min_col=1, max_col=1+NCOL, min_row=ROWMAP[key]), titles_from_data=True, from_rows=True)
bar.set_categories(cats)
line = LineChart()
line.add_data(Reference(ws, min_col=1, max_col=1+NCOL, min_row=ROWMAP["経常利益"]), titles_from_data=True, from_rows=True)
line.set_categories(cats)
bar += line
bar.title = "売上高（内訳）と経常利益の推移（百万円）"
bar.height = 9; bar.width = 22
bar.y_axis.numFmt = '#,##0'; bar.y_axis.delete = False; bar.x_axis.delete = False
chart_row = r + 1
ws.add_chart(bar, "A%d" % chart_row)

cf = BarChart(); cf.type = "col"; cf.grouping = "clustered"
for key in ["営業CF", "投資CF", "財務CF"]:
    cf.add_data(Reference(ws, min_col=1, max_col=1+NCOL, min_row=ROWMAP[key]), titles_from_data=True, from_rows=True)
cf.set_categories(cats)
cf.title = "キャッシュフローの推移（百万円）"
cf.height = 9; cf.width = 22
cf.y_axis.numFmt = '#,##0'; cf.y_axis.delete = False; cf.x_axis.delete = False
ws.add_chart(cf, "A%d" % (chart_row + 19))

out = "数値計画_損益全体像（短期）_整形版_2026-07-21.xlsx"
wb.save(out)
print("saved", out)
